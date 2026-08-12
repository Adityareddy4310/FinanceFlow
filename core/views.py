from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import PasswordResetView, PasswordResetConfirmView
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from datetime import datetime, timedelta, date
from django.urls import reverse_lazy
from django.core.mail import send_mail
from django.contrib import messages
from django.conf import settings
import json
import openpyxl
from openpyxl.styles import Font, PatternFill

from .models import (
    FinanceGroup, Borrower, WeeklyPayment, DailyExpense, DailyInterest,
    Employee, CollectionStaffEntry,
)
from .forms import CustomUserCreationForm
from django.db.models import Q, Prefetch, Sum, Min
from collections import defaultdict

def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')


def signup(request):
    """Signup with email and phone"""
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('dashboard')
    else:
        form = CustomUserCreationForm()

    return render(request, 'registration/signup.html', {'form': form})


class CustomPasswordResetView(PasswordResetView):
    """Password reset view with email"""
    template_name = 'registration/password_reset.html'
    email_template_name = 'registration/password_reset_email.html'
    subject_template_name = 'registration/password_reset_subject.txt'
    success_url = reverse_lazy('password_reset_done')

    def form_valid(self, form):
        """Send password reset email"""
        opts = {
            'use_https': self.request.is_secure(),
            'email_template_name': self.email_template_name,
            'subject_template_name': self.subject_template_name,
            'request': self.request,
            'html_email_template_name': None,
        }
        form.save(**opts)
        return super().form_valid(form)


def password_reset_done(request):
    """Password reset email sent confirmation"""
    return render(request, 'registration/password_reset_done.html')


class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    """Password reset confirm view"""
    template_name = 'registration/password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')


def password_reset_complete(request):
    """Password reset complete"""
    return render(request, 'registration/password_reset_complete.html')


@login_required
def dashboard(request):
    """Dashboard with finance groups"""
    groups = FinanceGroup.objects.filter(user=request.user).prefetch_related('borrowers')

    if not groups.exists():
        FinanceGroup.objects.create(user=request.user, name='Vizag Finance', day='Sunday Collection - Vizag', location='Vizag')
        FinanceGroup.objects.create(user=request.user, name='Eluru Finance', day='Monday Collection - Eluru', location='Eluru')
        groups = FinanceGroup.objects.filter(user=request.user).prefetch_related('borrowers')

    finance_groups = []
    for group in groups:
        active_borrowers = [b for b in group.borrowers.all() if b.name and not b.is_archived]
        total_balance = sum(float(b.balance) for b in active_borrowers)

        finance_groups.append({
            'id': group.id,
            'name': group.name,
            'day': group.day,
            'location': group.location,
            'active_borrowers': len(active_borrowers),
            'total_balance': int(total_balance),
        })

    return render(request, 'core/dashboard.html', {
        'finance_groups': finance_groups,
        'user': request.user,
    })


@login_required
@require_http_methods(["POST"])
def edit_finance_group(request, group_id):
    """Edit finance group name and day"""
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)

    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        day = data.get('day', '').strip()

        if name:
            group.name = name
        if day:
            group.day = day

        group.save()

        return JsonResponse({
            'success': True,
            'group': {
                'id': group.id,
                'name': group.name,
                'day': group.day,
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def group_detail(request, group_id):
    """Group detail with search support"""
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)

    # Get search query
    search_query = request.GET.get('search', '').strip()

    # Active (non-archived) borrowers only — this is the operational list.
    borrowers_query = Borrower.objects.filter(
        finance_group=group,
        name__gt='',
        is_archived=False,
    ).prefetch_related('weekly_payments')

    # Apply search filter
    if search_query:
        borrowers_query = borrowers_query.filter(
            Q(name__icontains=search_query) |
            Q(serial_number__icontains=search_query)
        )

    all_borrowers = list(borrowers_query.order_by('serial_number'))

    today = datetime.now().date()
    first_day = today.replace(day=1)

    if today.month == 12:
        last_day = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        last_day = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

    days = []
    current = first_day
    while current <= last_day:
        days.append({
            'date': current.isoformat(),
            'display': current.strftime('%d/%m')
        })
        current += timedelta(days=1)

    # Per-borrower day-grid: CURRENT loan_cycle only, so a new loan's cells
    # start blank even if the borrower closed a previous loan on the same
    # calendar date this month.
    borrower_data = []
    for borrower in all_borrowers:
        payments = {day['date']: 0 for day in days}
        for payment in borrower.weekly_payments.filter(loan_cycle=borrower.loan_cycle):
            payment_date_str = payment.payment_date.isoformat()
            if payment_date_str in payments:
                payments[payment_date_str] = float(payment.amount_paid)

        borrower_data.append({
            'id': borrower.id,
            'serial': borrower.serial_number,
            'name': borrower.name,
            'amount_given': float(borrower.amount_given),
            'amount_paid': float(borrower.amount_paid),
            'total_paid': borrower.total_paid,
            'balance': borrower.balance,
            'loan_date': borrower.date_of_loan.isoformat() if borrower.date_of_loan else '',
            'payments': payments
        })

    # Daily collection totals: EVERY WeeklyPayment row for the group this
    # month, across ALL loan cycles and INCLUDING archived borrowers — this
    # is the group's actual daily collection total and must not lose a
    # same-day old-loan payment or an archived borrower's history.
    daily_totals = {day['date']: 0 for day in days}
    all_payments_this_month = WeeklyPayment.objects.filter(
        borrower__finance_group=group,
        payment_date__gte=first_day,
        payment_date__lte=last_day,
    )
    for payment in all_payments_this_month:
        d = payment.payment_date.isoformat()
        if d in daily_totals:
            daily_totals[d] += float(payment.amount_paid)

    total_to_collect = sum(b['balance'] for b in borrower_data)
    month_name = today.strftime('%B %Y')

    return render(request, 'core/group_detail.html', {
        'group': group,
        'borrowers': borrower_data,
        'days': days,
        'daily_totals': daily_totals,
        'total_to_collect': total_to_collect,
        'month_name': month_name,
        'search_query': search_query,
    })


@login_required
def search_borrowers(request, group_id):
    """AJAX search endpoint for borrowers"""
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)
    search_query = request.GET.get('q', '').strip()

    if len(search_query) < 1:
        return JsonResponse({'results': []})

    borrowers = Borrower.objects.filter(
        finance_group=group,
        name__gt='',
        is_archived=False,
    ).filter(
        Q(name__icontains=search_query) |
        Q(serial_number__icontains=search_query)
    )[:10]  # Limit to 10 results

    results = []
    for borrower in borrowers:
        results.append({
            'id': borrower.id,
            'serial': borrower.serial_number,
            'name': borrower.name,
            'amount_given': float(borrower.amount_given),
            'balance': borrower.balance,
        })

    return JsonResponse({'results': results})


def _create_or_update_borrower(group, serial_number, name, amount_given, date_of_loan,
                                amount_paid=0, seed_month_payments=True, allow_overwrite=False):
    """
    Shared borrower creation/update logic.
    Used by both manual add_borrower (allow_overwrite=False, original behavior,
    blocks existing serials) and Excel import confirm (allow_overwrite=True only
    when the user explicitly chose "Replace" for a duplicate).
    Returns (borrower_or_None, created_bool, error_message_or_None).
    """
    existing = Borrower.objects.filter(finance_group=group, serial_number=serial_number).first()

    if existing and existing.name and not allow_overwrite:
        return None, False, f'Serial #{serial_number} already exists'

    was_existing = bool(existing and existing.name)

    if existing:
        borrower = existing
    else:
        borrower = Borrower(finance_group=group, serial_number=serial_number)

    borrower.name = name
    borrower.amount_given = amount_given
    borrower.amount_paid = amount_paid
    borrower.date_of_loan = date_of_loan
    borrower.save()

    if seed_month_payments:
        today = datetime.now().date()
        first_day = today.replace(day=1)

        if today.month == 12:
            last_day = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

        payments_to_create = []
        current = first_day
        while current <= last_day:
            payments_to_create.append(
                WeeklyPayment(borrower=borrower, payment_date=current, amount_paid=0, loan_cycle=borrower.loan_cycle)
            )
            current += timedelta(days=1)

        WeeklyPayment.objects.bulk_create(payments_to_create, ignore_conflicts=True)

    return borrower, (not was_existing), None


@login_required
@require_http_methods(["POST"])
def add_borrower(request, group_id):
    """Create new borrower (manual entry — unchanged behavior, now via shared helper)"""
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)

    try:
        data = json.loads(request.body)
        serial_number = int(data.get('serial_number', 0))
        name = data.get('name', '').strip()
        amount_given = float(data.get('amount_given', 0))
        date_of_loan_str = data.get('date_of_loan')

        if serial_number <= 0:
            return JsonResponse({'error': 'Serial number must be greater than 0'}, status=400)
        if not name:
            return JsonResponse({'error': 'Name is required'}, status=400)
        if amount_given <= 0:
            return JsonResponse({'error': 'Amount must be greater than 0'}, status=400)

        date_of_loan = datetime.fromisoformat(date_of_loan_str).date() if date_of_loan_str else date.today()

        borrower, created, error = _create_or_update_borrower(
            group, serial_number, name, amount_given, date_of_loan
        )

        if error:
            return JsonResponse({'error': error}, status=400)

        return JsonResponse({
            'success': True,
            'borrower': {
                'id': borrower.id,
                'serial': borrower.serial_number,
                'name': borrower.name,
            }
        })
    except Exception as e:
        return JsonResponse({'error': f'Error: {str(e)}'}, status=400)


@login_required
@require_http_methods(["POST"])
def delete_borrower(request, borrower_id):
    """
    Archive borrower (Blocker 1 fix). No longer a hard delete — the borrower
    disappears from active Borrower Records but their record and all
    WeeklyPayment history remain in the database permanently for historical
    cash-flow/reporting purposes.
    """
    borrower = get_object_or_404(Borrower, id=borrower_id, finance_group__user=request.user)
    borrower.is_archived = True
    borrower.save()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def update_payment(request, borrower_id):
    """Update payment"""
    borrower = get_object_or_404(Borrower, id=borrower_id, finance_group__user=request.user)

    try:
        data = json.loads(request.body)
        payment_date_str = data.get('date')
        amount = float(data.get('amount', 0))

        payment_date = datetime.fromisoformat(payment_date_str).date()

        # Stamped with the borrower's CURRENT loan_cycle at write time, so
        # this payment is permanently tied to whichever loan was active when
        # it was actually collected.
        payment, created = WeeklyPayment.objects.get_or_create(
            borrower=borrower,
            payment_date=payment_date,
            loan_cycle=borrower.loan_cycle,
            defaults={'amount_paid': amount}
        )

        if not created:
            payment.amount_paid = amount
            payment.save()

        return JsonResponse({
            'success': True,
            'total_paid': borrower.total_paid,
            'balance': borrower.balance
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def update_amount_paid(request, borrower_id):
    """Update amount paid"""
    borrower = get_object_or_404(Borrower, id=borrower_id, finance_group__user=request.user)

    try:
        data = json.loads(request.body)
        amount = float(data.get('amount', 0))

        borrower.amount_paid = amount
        borrower.save()

        return JsonResponse({
            'success': True,
            'total_paid': borrower.total_paid,
            'balance': borrower.balance
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def update_borrower(request, borrower_id):
    """Update borrower details"""
    borrower = get_object_or_404(Borrower, id=borrower_id, finance_group__user=request.user)

    try:
        data = json.loads(request.body)

        if 'name' in data:
            borrower.name = data['name'].strip()

        if 'amount_given' in data:
            borrower.amount_given = float(data['amount_given'])

        if 'date_of_loan' in data and data['date_of_loan']:
            borrower.date_of_loan = datetime.fromisoformat(data['date_of_loan']).date()

        borrower.save()

        return JsonResponse({
            'success': True,
            'borrower': {
                'id': borrower.id,
                'name': borrower.name,
                'amount_given': float(borrower.amount_given),
                'total_paid': borrower.total_paid,
                'balance': borrower.balance
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ============================================================
# EXCEL IMPORT / EXPORT — new in this phase
# ============================================================

def _parse_excel_row(row_cells, row_index):
    """
    Validate a single row from the uploaded workbook.
    Expected columns (in order): Serial Number, Borrower Name, Loan Date, Loan Amount
    Returns (row_data_dict_or_None, error_string_or_None).
    row_data is None only for a fully empty row (silently skipped).
    """
    padded = (list(row_cells) + [None, None, None, None])[:4]
    serial_raw, name_raw, date_raw, amount_raw = padded

    if all(c is None or str(c).strip() == '' for c in padded):
        return None, None  # fully empty row

    errors = []

    try:
        serial_number = int(float(serial_raw))
        if serial_number <= 0:
            errors.append('Serial number must be greater than 0')
    except (TypeError, ValueError):
        serial_number = None
        errors.append('Missing or invalid serial number')

    name = str(name_raw).strip() if name_raw is not None else ''
    if not name:
        errors.append('Missing borrower name')

    date_of_loan = None
    if date_raw is None or str(date_raw).strip() == '':
        errors.append('Missing loan date')
    else:
        if isinstance(date_raw, datetime):
            date_of_loan = date_raw.date()
        elif isinstance(date_raw, date):
            date_of_loan = date_raw
        else:
            date_str = str(date_raw).strip()
            parsed = None
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    parsed = datetime.strptime(date_str, fmt).date()
                    break
                except ValueError:
                    continue
            if parsed is None:
                errors.append('Invalid date format (use DD/MM/YYYY)')
            else:
                date_of_loan = parsed

    try:
        amount_given = float(amount_raw)
        if amount_given <= 0:
            errors.append('Loan amount must be greater than 0')
    except (TypeError, ValueError):
        amount_given = None
        errors.append('Missing or invalid loan amount')

    row_data = {
        'row': row_index,
        'serial_number': serial_number,
        'name': name,
        'date_of_loan': date_of_loan.isoformat() if date_of_loan else None,
        'amount_given': amount_given,
    }

    if errors:
        return row_data, '; '.join(errors)

    return row_data, None


@login_required
@require_http_methods(["POST"])
def import_borrowers_preview(request, group_id):
    """
    Parse an uploaded Excel file and return a validated preview.
    Writes nothing to the database — confirmation happens in import_borrowers_confirm.
    """
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)

    excel_file = request.FILES.get('file')
    if not excel_file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
        return JsonResponse({'error': 'File must be .xlsx or .xls'}, status=400)

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
    except Exception:
        return JsonResponse({'error': 'Could not read the Excel file. Please check the format.'}, status=400)

    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # row 1 assumed header
    if not rows:
        return JsonResponse({'error': 'No data rows found in the file'}, status=400)

    existing_serials = set(
        Borrower.objects.filter(finance_group=group, name__gt='').values_list('serial_number', flat=True)
    )

    valid_rows = []
    error_rows = []
    seen_serials_in_file = set()

    for idx, row in enumerate(rows, start=2):
        parsed, error = _parse_excel_row(list(row), idx)
        if parsed is None:
            continue  # empty row, skip silently

        if error:
            error_rows.append({**parsed, 'error': error})
            continue

        if parsed['serial_number'] in seen_serials_in_file:
            error_rows.append({**parsed, 'error': f'Duplicate serial number within file (row {idx})'})
            continue
        seen_serials_in_file.add(parsed['serial_number'])

        parsed['is_duplicate'] = parsed['serial_number'] in existing_serials
        valid_rows.append(parsed)

    return JsonResponse({
        'success': True,
        'total_detected': len(valid_rows) + len(error_rows),
        'valid_count': len(valid_rows),
        'error_count': len(error_rows),
        'duplicate_count': sum(1 for r in valid_rows if r['is_duplicate']),
        'valid_rows': valid_rows,
        'error_rows': error_rows,
    })


@login_required
@require_http_methods(["POST"])
def import_borrowers_confirm(request, group_id):
    """
    Commit previously-previewed rows to the database.
    Wrapped in a single atomic transaction — reuses the same borrower creation
    helper as manual add, so imported borrowers behave identically (same
    WeeklyPayment seeding, same calculations).
    """
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)

    try:
        data = json.loads(request.body)
        rows = data.get('rows', [])
        duplicate_action = data.get('duplicate_action', 'skip')

        if duplicate_action not in ('skip', 'replace'):
            return JsonResponse({'error': 'Invalid duplicate action'}, status=400)

        created_count = 0
        updated_count = 0
        skipped_count = 0
        row_errors = []

        with transaction.atomic():
            for row in rows:
                try:
                    serial_number = int(row.get('serial_number'))
                    name = str(row.get('name', '')).strip()
                    amount_given = float(row.get('amount_given'))
                    date_of_loan = datetime.fromisoformat(row.get('date_of_loan')).date()
                except (TypeError, ValueError):
                    row_errors.append(f"Row with serial {row.get('serial_number')}: invalid data, skipped")
                    continue

                if serial_number <= 0 or not name or amount_given <= 0:
                    row_errors.append(f'Row with serial {serial_number}: failed validation, skipped')
                    continue

                is_duplicate = Borrower.objects.filter(
                    finance_group=group, serial_number=serial_number, name__gt=''
                ).exists()

                if is_duplicate and duplicate_action == 'skip':
                    skipped_count += 1
                    continue

                borrower, created, error = _create_or_update_borrower(
                    group, serial_number, name, amount_given, date_of_loan,
                    allow_overwrite=(duplicate_action == 'replace')
                )

                if error:
                    row_errors.append(f'Row with serial {serial_number}: {error}')
                    continue

                if is_duplicate:
                    updated_count += 1
                else:
                    created_count += 1

        return JsonResponse({
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'row_errors': row_errors,
        })
    except Exception as e:
        return JsonResponse({'error': f'Import failed: {str(e)}'}, status=400)


@login_required
def export_borrowers_excel(request, group_id):
    """
    Generate a real .xlsx backup of all borrowers in this group.
    First 4 columns match the import format exactly, so an exported file
    can be re-imported with minimal effort.
    """
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)
    borrowers = Borrower.objects.filter(finance_group=group, name__gt='', is_archived=False).order_by('serial_number')

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Borrowers'

    headers = ['Serial Number', 'Borrower Name', 'Loan Date', 'Loan Amount',
               'Amount Paid', 'Outstanding Balance', 'Status']
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')

    for b in borrowers:
        status = 'Cleared' if b.balance <= 0 else 'Active'
        sheet.append([
            b.serial_number,
            b.name,
            b.date_of_loan.strftime('%d/%m/%Y') if b.date_of_loan else '',
            float(b.amount_given),
            float(b.total_paid),
            float(b.balance),
            status,
        ])

    for col_cells in sheet.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        sheet.column_dimensions[col_cells[0].column_letter].width = max(12, length + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{group.name.replace(' ', '_')}_borrowers_backup.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@require_http_methods(["POST"])
def give_new_loan(request, borrower_id):
    """
    Start a new loan cycle for a borrower (typically one who has fully repaid).
    Increments loan_cycle and resets amount_paid/date_of_loan. Does NOT touch
    any existing WeeklyPayment row — the old cycle's payments (including a
    same-day final payment that closed the previous loan) remain in the
    database exactly as recorded, permanently, and keep counting toward the
    group's daily collection total. Borrower.total_paid now excludes them
    automatically because it filters by the (now incremented) loan_cycle.
    """
    borrower = get_object_or_404(Borrower, id=borrower_id, finance_group__user=request.user)
    try:
        data = json.loads(request.body)
        new_amount = float(data.get('amount_given', 0))
        new_date_str = data.get('date_of_loan')

        if new_amount <= 0:
            return JsonResponse({'error': 'Amount must be greater than 0'}, status=400)

        new_date = datetime.fromisoformat(new_date_str).date() if new_date_str else date.today()

        borrower.amount_given = new_amount
        borrower.amount_paid = 0
        borrower.date_of_loan = new_date
        borrower.loan_cycle = borrower.loan_cycle + 1
        borrower.save()

        return JsonResponse({
            'success': True,
            'borrower': {
                'id': borrower.id,
                'amount_given': float(borrower.amount_given),
                'date_of_loan': borrower.date_of_loan.isoformat(),
                'total_paid': borrower.total_paid,
                'balance': borrower.balance,
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def cash_flow_data(request, group_id):
    """Read-only daily cash flow, derived entirely from existing WeeklyPayment/Borrower data.
    NOTE: shadowed by cash_flow_summary at the same URL in urls.py (both map to
    api/group/<id>/cash-flow/, cash_flow_summary is registered second and wins).
    Left as-is — previously flagged, not touched, since urls.py still imports it."""
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)

    earliest_loan = Borrower.objects.filter(finance_group=group, name__gt='').aggregate(Min('date_of_loan'))['date_of_loan__min']
    earliest_payment = WeeklyPayment.objects.filter(borrower__finance_group=group).aggregate(Min('payment_date'))['payment_date__min']
    candidates = [d for d in [earliest_loan, earliest_payment] if d]

    if not candidates:
        return JsonResponse({'success': True, 'entries': [], 'summary': {
            'today_collected': 0, 'today_loans': 0, 'today_net': 0, 'current_available_cash': 0
        }})

    start_date = min(candidates)
    today = date.today()

    collected_by_date = defaultdict(float)
    for row in WeeklyPayment.objects.filter(
        borrower__finance_group=group, payment_date__gte=start_date, payment_date__lte=today
    ).values('payment_date').annotate(total=Sum('amount_paid')):
        collected_by_date[row['payment_date']] = float(row['total'] or 0)

    loans_by_date = defaultdict(float)
    for row in Borrower.objects.filter(
        finance_group=group, name__gt='', date_of_loan__gte=start_date, date_of_loan__lte=today
    ).values('date_of_loan').annotate(total=Sum('amount_given')):
        loans_by_date[row['date_of_loan']] = float(row['total'] or 0)

    entries = []
    running = 0.0
    current = start_date
    while current <= today:
        collected = collected_by_date.get(current, 0.0)
        loans = loans_by_date.get(current, 0.0)
        net = collected - loans
        running += net
        entries.append({
            'date': current.isoformat(),
            'display': current.strftime('%d %b'),
            'collected': collected,
            'loans_issued': loans,
            'net_change': net,
            'running_balance': running,
        })
        current += timedelta(days=1)

    latest = entries[-1] if entries else None
    summary = {
        'today_collected': latest['collected'] if latest else 0,
        'today_loans': latest['loans_issued'] if latest else 0,
        'today_net': latest['net_change'] if latest else 0,
        'current_available_cash': latest['running_balance'] if latest else 0,
    }

    return JsonResponse({'success': True, 'entries': entries, 'summary': summary})


@login_required
def cash_flow_summary(request, group_id):
    """
    Daily cash flow ledger. Read-only aggregation over existing data.
    Collections: WeeklyPayment grouped by payment_date, across all borrowers in group.
    Loans issued: Borrower.amount_given grouped by date_of_loan.
    NOTE: this is the one actually reachable at api/group/<id>/cash-flow/ (registered
    second in urls.py). group_detail.html doesn't call it — Cash Flow drawer computes
    client-side. Left as-is, previously flagged."""
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)

    payments = WeeklyPayment.objects.filter(
        borrower__finance_group=group
    ).values('payment_date').annotate(total=Sum('amount_paid')).order_by('payment_date')
    collected_by_date = {p['payment_date'].isoformat(): float(p['total']) for p in payments if p['total']}

    loans = Borrower.objects.filter(
        finance_group=group, name__gt='', date_of_loan__isnull=False
    ).values('date_of_loan').annotate(total=Sum('amount_given')).order_by('date_of_loan')
    loaned_by_date = {l['date_of_loan'].isoformat(): float(l['total']) for l in loans if l['total']}

    all_dates = sorted(set(collected_by_date.keys()) | set(loaned_by_date.keys()))

    ledger = []
    running_balance = 0
    for d in all_dates:
        collected = collected_by_date.get(d, 0)
        loaned = loaned_by_date.get(d, 0)
        net = collected - loaned
        running_balance += net
        ledger.append({'date': d, 'collected': collected, 'loaned': loaned, 'net': net, 'running_balance': running_balance})

    today_str = datetime.now().date().isoformat()
    today_entry = next((row for row in ledger if row['date'] == today_str), None)

    return JsonResponse({
        'success': True,
        'ledger': ledger,
        'today': {
            'collected': today_entry['collected'] if today_entry else 0,
            'loaned': today_entry['loaned'] if today_entry else 0,
            'net': today_entry['net'] if today_entry else 0,
        },
        'current_available_cash': running_balance,
    })


# ============================================================
# CASH FLOW — Interest, Expenses, Collection Staff (date-selectable)
# ============================================================

@login_required
def cash_flow_extras(request, group_id):
    """
    Returns expenses + interest + collection staff for the group's current
    month, plus the full active employee list (for populating the
    multi-select). Auto-seeds 3 default employees on first use per account,
    same pattern as dashboard()'s auto-seed of default groups.
    """
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)
    today = date.today()
    first_day = today.replace(day=1)

    if not Employee.objects.filter(user=request.user).exists():
        Employee.objects.bulk_create([
            Employee(user=request.user, name='Srikanth'),
            Employee(user=request.user, name='Rama Reddy'),
            Employee(user=request.user, name='Adi'),
        ])

    expenses = DailyExpense.objects.filter(finance_group=group, date__gte=first_day, date__lte=today)
    expenses_by_date = defaultdict(list)
    for e in expenses:
        expenses_by_date[e.date.isoformat()].append({'id': e.id, 'category': e.category, 'amount': float(e.amount)})

    interest = DailyInterest.objects.filter(finance_group=group, date__gte=first_day, date__lte=today)
    interest_by_date = {i.date.isoformat(): float(i.amount) for i in interest}

    staff_entries = CollectionStaffEntry.objects.filter(
        finance_group=group, date__gte=first_day, date__lte=today
    ).select_related('employee')
    staff_by_date = defaultdict(list)
    for s in staff_entries:
        staff_by_date[s.date.isoformat()].append({'id': s.employee_id, 'name': s.employee.name})

    employees = list(Employee.objects.filter(user=request.user, is_active=True).values('id', 'name'))

    return JsonResponse({
        'success': True,
        'expenses': expenses_by_date,
        'interest': interest_by_date,
        'collection_staff': staff_by_date,
        'employees': employees,
    })


@login_required
@require_http_methods(["POST"])
def add_expense(request, group_id):
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)
    try:
        data = json.loads(request.body)
        exp_date = datetime.fromisoformat(data.get('date')).date()
        category = data.get('category')
        amount = float(data.get('amount', 0))
        valid_categories = {'petrol', 'food', 'room_rent', 'salaries', 'misc'}
        if category not in valid_categories or amount <= 0:
            return JsonResponse({'error': 'Invalid category or amount'}, status=400)
        expense = DailyExpense.objects.create(finance_group=group, date=exp_date, category=category, amount=amount)
        return JsonResponse({'success': True, 'expense': {'id': expense.id, 'category': category, 'amount': amount}})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def delete_expense(request, expense_id):
    expense = get_object_or_404(DailyExpense, id=expense_id, finance_group__user=request.user)
    expense.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def update_interest(request, group_id):
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)
    try:
        data = json.loads(request.body)
        int_date = datetime.fromisoformat(data.get('date')).date()
        amount = float(data.get('amount', 0))
        obj, _ = DailyInterest.objects.update_or_create(
            finance_group=group, date=int_date, defaults={'amount': amount}
        )
        return JsonResponse({'success': True, 'amount': float(obj.amount)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def save_collection_staff(request, group_id):
    """
    Replace-all-for-date: given a date and a list of employee_ids, sets the
    collection staff for that (group, date) to exactly that set. Informational
    only — never touches Borrower/WeeklyPayment/DailyExpense/DailyInterest or
    any cash-flow calculation.
    """
    group = get_object_or_404(FinanceGroup, id=group_id, user=request.user)
    try:
        data = json.loads(request.body)
        entry_date = datetime.fromisoformat(data.get('date')).date()
        employee_ids = data.get('employee_ids', [])

        valid_employee_ids = set(
            Employee.objects.filter(user=request.user, id__in=employee_ids).values_list('id', flat=True)
        )

        with transaction.atomic():
            CollectionStaffEntry.objects.filter(finance_group=group, date=entry_date).delete()
            CollectionStaffEntry.objects.bulk_create([
                CollectionStaffEntry(finance_group=group, date=entry_date, employee_id=eid)
                for eid in valid_employee_ids
            ])

        names = list(
            Employee.objects.filter(id__in=valid_employee_ids).values_list('name', flat=True)
        )
        return JsonResponse({'success': True, 'staff': names})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def contact(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        message_text = request.POST.get('message', '').strip()

        if name and email and message_text:
            try:
                send_mail(
                    subject=f'FinanceFlow Contact: {name}',
                    message=f'From: {name} <{email}>\n\n{message_text}',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[settings.EMAIL_HOST_USER],
                    fail_silently=False,
                )
                messages.success(request, 'Message sent successfully.')
            except Exception:
                messages.error(request, 'Could not send message. Please try again later.')
        else:
            messages.error(request, 'Please fill all fields.')
        return redirect('contact')

    return render(request, 'core/contact.html')